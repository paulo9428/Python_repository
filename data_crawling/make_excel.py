import openpyxl
import datetime


book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"
sheet1.cell(row=1, column=1).value = 'Title'
sheet2 = book.create_sheet()
sheet2.title = "두번째 시트"
sheet2['A1'] = datetime.datetime.now()
sheet2['A2'] = datetime.date.today()
# 저장하기
book.save("./excelfile/output.xlsx")
