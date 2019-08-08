import requests
from bs4 import BeautifulSoup
import openpyxl
import csv, codecs
from PIL import Image
from openpyxl.chart import (
    Reference, Series,
    BarChart, ScatterChart
)

book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "멜론 TOP100"

with codecs.open('./meltop100.csv', 'r', 'utf-8') as meltop:
    reader = csv.reader(meltop, delimiter=',', quotechar='"')

    for i, row in enumerate(reader):
   # print(i, row)
        for j, col in enumerate(row):
            tcell = sheet1.cell(row=(i+1), column=j+1)
            if i > 0 and (j == 0 or j > 2) and col.isnumeric():
                print(j, col)
                tcell.number_format
                tcell.value = int(col)
            else:
                tcell.value = col

##------------------------------------------------------------------------------------

sheet2 = book.create_sheet()
sheet2.title = "앨범 자켓 이미지"

for i in range(1, 101):
 
    imgFile = 'C:\workspace\Learn_Python\crawl\image\{}.jpg'.format(i)

    img2 = Image.open(imgFile)
    new_img = img2.resize((30, 30))
    new_img.save('new{}.png'.format(i))
    img3 = openpyxl.drawing.image.Image('new{}.png'.format(i))
    sheet2.add_image(img3, 'A{}'.format(i))

##---------------------------------------------------------------------------------------------

sheet3 = book.create_sheet()
sheet3.title = "좋아요 수 차트"

datax = Reference(sheet1, min_col=4,               
		min_row=2, max_col=4, max_row=11)
categs = Reference(sheet1, min_col=2, 
				 min_row=2, max_row=11)

chart = BarChart()
chart.add_data(data=datax)
chart.set_categories(categs)

chart.legend = None  # 범례
chart.varyColors = True
chart.title = "좋아요 차트"

sheet3.add_chart(chart, "A3")



##-------------------------------------------------------------------------------------------


chart = ScatterChart()
chart.style = 13
chart.x_axis.title = '노래'
chart.y_axis.title = '좋아요차이 수'

xvalues = Reference(sheet1, min_col=2,
max_col=2, min_row=2, max_row=11)


values = Reference(sheet1, 
            min_col=5, 
            min_row=2, 
            max_row=11)


series = Series(values, xvalues)
chart.series.append(series)

sheet3.add_chart(chart, "A10")


book.save("./output3.xlsx")




















