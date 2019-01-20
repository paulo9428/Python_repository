import openpyxl
import csv, codecs


book = openpyxl.Workbook()
sheet1 = book.active

sheet1.title = "멜론 TOP100"



# top_lst = []

# fp = codecs.open("./meltop100.csv", "r", "utf-8")

# reader = csv.reader(fp, delimiter=',', quotechar='"')

# # with codecs.open('./output.csv', 'w', 'utf-8') as ff:
# #     writer = csv.writer(ff, delimiter=',', quotechar='"')

# for cells in reader:
#     writer.writerow([cells[0], random.randrange(1,100)])    


with codecs.open('./meltop100.csv', 'r', 'utf-8') as meltop:
    reader = csv.reader(meltop, delimiter=',', quotechar='"')
    for i, row in enumerate(reader):
        # print(row)
        
        for j in range(0, 5):

            sheet1.cell(row=i+1, column=j+1).value = row[j]

for i in range(2, 102):

    tmpCell_D = sheet1['D{}'.format(i)]
    tmpCell_D.number_format

    temCell_E = sheet1['E{}'.format(i)]
    tmpCell_E.number_format

## int() 사용하세요 Slack 참고

# with open('./meltop100.csv', mode = 'r', encoding = 'utf-8') as meltop:
#     for line in meltop:
#         top_lst.append(line.split(','))

# # print(top_lst)

# for i in range(1, 102):
#     for j in range(1, 5):
#         sheet1.cell(row=i, column=j).value = top_lst[i][j]


book.save("./output3.xlsx")



